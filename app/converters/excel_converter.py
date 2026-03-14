"""Excel 导出器 — 将表格块导出为 xlsx，非表格内容降级到第一个 sheet。"""

from __future__ import annotations

from pathlib import Path

from app.converters.base_converter import BaseConverter
from app.models import DocumentResult, BlockType


class ExcelConverter(BaseConverter):
    @property
    def file_extension(self) -> str:
        return ".xlsx"

    def convert(self, result: DocumentResult, output_path: Path) -> Path:
        from openpyxl import Workbook

        wb = Workbook()
        # 删除默认 sheet，后面按需创建
        wb.remove(wb.active)

        table_count = 0

        for page in result.pages:
            for block in page.blocks:
                if block.block_type == BlockType.TABLE and block.table_cells:
                    table_count += 1
                    ws = wb.create_sheet(title=f"Table_{table_count}")
                    for row_idx, row in enumerate(block.table_cells, 1):
                        for col_idx, cell in enumerate(row, 1):
                            ws.cell(row=row_idx, column=col_idx, value=str(cell))

        # 降级：无表格时，将纯文本逐行写入
        if table_count == 0:
            ws = wb.create_sheet(title="OCR Text")
            text = result.plain_text or ""
            for row_idx, line in enumerate(text.split("\n"), 1):
                if line.strip():
                    ws.cell(row=row_idx, column=1, value=line)

        wb.save(str(output_path))
        return output_path
