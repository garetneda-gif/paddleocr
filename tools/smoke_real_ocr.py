"""真实 OCR + 导出烟测脚本。"""

from __future__ import annotations

import json
from pathlib import Path
import sys
from tempfile import TemporaryDirectory

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from app.core.export_router import create_default_router
from app.core.onnx_engine import OnnxOCREngine
from app.models.enums import OutputFormat


def main() -> None:
    image_path = Path("tests/fixtures/test_en.png").resolve()
    doc = OnnxOCREngine(lang="en", speed_mode="mobile").predict(image_path)
    router = create_default_router()
    summary: dict[str, object] = {
        "plain_text": doc.plain_text[:200],
        "page_count": doc.page_count,
        "block_count": len(doc.pages[0].blocks) if doc.pages else 0,
        "outputs": {},
    }

    with TemporaryDirectory(prefix="paddleocr_exports_") as tmp:
        out_dir = Path(tmp)
        for fmt in OutputFormat:
            converter = router.select_converter(fmt)
            output_path = out_dir / f"test_en{converter.file_extension}"
            converter.convert(doc, output_path)

            item: dict[str, object] = {
                "path": str(output_path),
                "size": output_path.stat().st_size,
            }
            if fmt in (OutputFormat.TXT, OutputFormat.HTML, OutputFormat.RTF):
                item["snippet"] = output_path.read_text(encoding="utf-8")[:160]
            elif fmt is OutputFormat.EXCEL:
                item["sheets"] = load_workbook(output_path).sheetnames

            summary["outputs"][fmt.value] = item

        print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
