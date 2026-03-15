"""导出器单元测试 — 使用构造的 DocumentResult，不依赖 PaddleOCR。"""

import sys
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).parent.parent))

from app.models import BlockResult, BlockType, DocumentResult, PageResult
from app.converters.txt_converter import TxtConverter
from app.converters.rtf_converter import RtfConverter
from app.converters.word_converter import WordConverter
from app.converters.html_converter import HtmlConverter
from app.converters.excel_converter import ExcelConverter


@pytest.fixture
def sample_result() -> DocumentResult:
    return DocumentResult(
        source_path=Path("/tmp/test.png"),
        page_count=1,
        pages=[
            PageResult(
                page_index=0,
                width=800,
                height=400,
                blocks=[
                    BlockResult(
                        block_type=BlockType.TITLE,
                        bbox=(10.0, 10.0, 200.0, 40.0),
                        text="Test Title",
                        confidence=0.99,
                    ),
                    BlockResult(
                        block_type=BlockType.PARAGRAPH,
                        bbox=(10.0, 50.0, 400.0, 80.0),
                        text="This is a paragraph of text.",
                        confidence=0.95,
                    ),
                    BlockResult(
                        block_type=BlockType.TABLE,
                        bbox=(10.0, 100.0, 400.0, 200.0),
                        text="",
                        table_cells=[
                            ["Name", "Age", "Score"],
                            ["Alice", "25", "98.5"],
                            ["Bob", "30", "87.0"],
                        ],
                    ),
                ],
            )
        ],
        plain_text="Test Title\nThis is a paragraph of text.",
    )


class TestTxtConverter:
    def test_creates_file(self, sample_result, tmp_path):
        out = tmp_path / "out.txt"
        TxtConverter().convert(sample_result, out)
        assert out.exists()
        content = out.read_text()
        assert "Test Title" in content
        assert "paragraph" in content


class TestRtfConverter:
    def test_creates_file(self, sample_result, tmp_path):
        out = tmp_path / "out.rtf"
        RtfConverter().convert(sample_result, out)
        assert out.exists()
        content = out.read_text()
        assert r"\rtf1" in content
        assert "Test Title" in content

    def test_fallback_plain_text(self, tmp_path):
        result = DocumentResult(
            source_path=Path("/tmp/test.pdf"),
            page_count=1,
            pages=[],
            plain_text="Fallback line 1\nFallback line 2",
        )
        out = tmp_path / "fallback.rtf"
        RtfConverter().convert(result, out)
        content = out.read_text()
        assert "Fallback line 1" in content
        assert "Fallback line 2" in content


class TestWordConverter:
    def test_creates_file(self, sample_result, tmp_path):
        out = tmp_path / "out.docx"
        WordConverter().convert(sample_result, out)
        assert out.exists()
        assert out.stat().st_size > 0

    def test_fallback_plain_text(self, tmp_path):
        result = DocumentResult(
            source_path=Path("/tmp/test.png"),
            page_count=1,
            pages=[PageResult(page_index=0, width=100, height=100, blocks=[])],
            plain_text="Fallback text line",
        )
        out = tmp_path / "fallback.docx"
        WordConverter().convert(result, out)
        assert out.exists()


class TestHtmlConverter:
    def test_creates_file(self, sample_result, tmp_path):
        out = tmp_path / "out.html"
        HtmlConverter().convert(sample_result, out)
        assert out.exists()
        content = out.read_text()
        assert "<h1>" in content
        assert "Test Title" in content
        assert "<table>" in content

    def test_escapes_html(self, tmp_path):
        result = DocumentResult(
            source_path=Path("/tmp/test.png"),
            page_count=1,
            pages=[
                PageResult(
                    page_index=0, width=100, height=100,
                    blocks=[
                        BlockResult(
                            block_type=BlockType.PARAGRAPH,
                            bbox=(0, 0, 100, 20),
                            text="<script>alert('xss')</script>",
                        )
                    ],
                )
            ],
            plain_text="",
        )
        out = tmp_path / "xss.html"
        HtmlConverter().convert(result, out)
        content = out.read_text()
        assert "<script>" not in content
        assert "&lt;script&gt;" in content


class TestExcelConverter:
    def test_creates_file_with_table(self, sample_result, tmp_path):
        out = tmp_path / "out.xlsx"
        ExcelConverter().convert(sample_result, out)
        assert out.exists()

        from openpyxl import load_workbook
        wb = load_workbook(str(out))
        assert "Table_1" in wb.sheetnames
        ws = wb["Table_1"]
        assert ws.cell(1, 1).value == "Name"
        assert ws.cell(2, 1).value == "Alice"

    def test_fallback_no_tables(self, tmp_path):
        result = DocumentResult(
            source_path=Path("/tmp/test.png"),
            page_count=1,
            pages=[
                PageResult(
                    page_index=0, width=100, height=100,
                    blocks=[
                        BlockResult(
                            block_type=BlockType.PARAGRAPH,
                            bbox=(0, 0, 100, 20),
                            text="Just text",
                        )
                    ],
                )
            ],
            plain_text="Just text",
        )
        out = tmp_path / "fallback.xlsx"
        ExcelConverter().convert(result, out)
        assert out.exists()

        from openpyxl import load_workbook
        wb = load_workbook(str(out))
        assert "OCR Text" in wb.sheetnames
